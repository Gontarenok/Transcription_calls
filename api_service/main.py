from __future__ import annotations

import os
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import urlencode

from fastapi import Depends, FastAPI, Form, HTTPException, Query, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import load_only, selectinload
from starlette.middleware.sessions import SessionMiddleware

from api_service.config import settings
from api_service.observability import (
    RequestContextMiddleware,
    instrument_prometheus,
    maybe_setup_opentelemetry,
    setup_logging,
)
from api_service.auth import (
    ROLE_911,
    ROLE_ADMIN,
    ROLE_KC,
    UiIdentity,
    allowed_call_types_for_role,
    get_current_role,
    get_current_identity_ui,
    identity_sees_911_summarization,
    identity_sees_kc_classification,
    menu_context,
    ui_login_authenticate,
)
from api_service.schemas import CallOut, CallsResponse, PipelineRunOut, PipelineRunsResponse, TopicCatalogEntriesResponse, TopicCatalogEntryOut, UserOut, UsersResponse
from jobs.generate_synonyms import catalog_generate_synonyms
from db.base import SessionLocal
from db.crud import (
    count_calls_with_active_classification,
    count_calls_with_latest_summarization_911,
    list_calls_with_active_classification,
    list_calls_with_latest_summarization_911,
    list_distinct_latest_summary_outcomes_911,
    list_distinct_latest_summary_topics_911,
    list_manager_names_for_summarized_911_filters,
    list_topic_catalog_entries,
    set_catalog_qdrant_point_id,
    update_topic_catalog_entry,
    upsert_topic_catalog_entry,
)
from db.models import Call, CallClassification, CallStatus, CallType, PipelineRun, TopicCatalogEntry, Transcription, User
from classification_rag.catalog_service import entry_source_hash, sync_catalog_entries

setup_logging()

app = FastAPI(title="Audio Calls API", version="0.6.0")
templates = Jinja2Templates(directory="api_service/templates")

_static_dir = Path(__file__).resolve().parent / "static"
if _static_dir.is_dir():
    app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")

app.add_middleware(
    SessionMiddleware,
    secret_key=settings.session_secret,
    same_site="lax",
    https_only=settings.session_cookie_secure,
)
app.add_middleware(RequestContextMiddleware)

instrument_prometheus(app)
maybe_setup_opentelemetry(app)


@app.get("/healthz")
def healthz():
    return {"status": "ok"}

STATUS_OPTIONS = [
    "NEW",
    "TRANSCRIBING",
    "TRANSCRIBED",
    "TRANSCRIPTION_FAILED",
    "CLASSIFYING",
    "CLASSIFIED",
    "CLASSIFICATION_FAILED",
    "SUMMARIZING",
    "SUMMARIZED",
    "SUMMARIZATION_FAILED",
    "FAILED",
]

# UI: размер страницы списков «Звонки» / «Классификация» (меньше страница — быстрее первая отрисовка; полный объём — Excel / API)
_UI_PAGE_SIZE = 200
# Если по фильтру больше записей — показываем предупреждение (мягкий порог)
_UI_LIST_SOFT_CAP = 50_000
# JSON API `/api/calls`: максимальный limit в одном запросе
_API_LIMIT_MAX = 2_000_000


def _calls_list_querystring(
    *,
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    status: str | None,
    call_type: str | None,
    offset: int,
) -> str:
    q: dict[str, str] = {}
    if offset > 0:
        q["offset"] = str(offset)
    if period:
        q["period"] = period
    if date_from:
        q["date_from"] = date_from
    if date_to:
        q["date_to"] = date_to
    if manager:
        q["manager"] = manager
    if status:
        q["status"] = status
    ct = normalize_call_type_filter(call_type) or (call_type or "").strip()
    if ct:
        q["call_type"] = ct
    return urlencode(q)


def _classified_list_querystring(
    *,
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    topic: str | None,
    subtopic: str | None,
    offset: int,
) -> str:
    q: dict[str, str] = {}
    if offset > 0:
        q["offset"] = str(offset)
    if period:
        q["period"] = period
    if date_from:
        q["date_from"] = date_from
    if date_to:
        q["date_to"] = date_to
    if manager:
        q["manager"] = manager
    if topic:
        q["topic"] = topic
    if subtopic:
        q["subtopic"] = subtopic
    return urlencode(q)


def _summarized_list_querystring(
    *,
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    topic: str | None,
    outcome: str | None,
    offset: int,
) -> str:
    q: dict[str, str] = {}
    if offset > 0:
        q["offset"] = str(offset)
    if period:
        q["period"] = period
    if date_from:
        q["date_from"] = date_from
    if date_to:
        q["date_to"] = date_to
    if manager:
        q["manager"] = manager
    if topic:
        q["topic"] = topic
    if outcome:
        q["outcome"] = outcome
    return urlencode(q)


def resolve_period(period: str | None) -> tuple[datetime | None, datetime | None]:
    if not period:
        return None, None
    today = date.today()
    if period == "today":
        return datetime.combine(today, time.min).replace(tzinfo=timezone.utc), datetime.combine(today, time.max).replace(tzinfo=timezone.utc)
    if period == "yesterday":
        d = today - timedelta(days=1)
        return datetime.combine(d, time.min).replace(tzinfo=timezone.utc), datetime.combine(d, time.max).replace(tzinfo=timezone.utc)
    if period == "week_current":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return datetime.combine(start, time.min).replace(tzinfo=timezone.utc), datetime.combine(end, time.max).replace(tzinfo=timezone.utc)
    if period == "week_prev":
        end = today - timedelta(days=today.weekday() + 1)
        start = end - timedelta(days=6)
        return datetime.combine(start, time.min).replace(tzinfo=timezone.utc), datetime.combine(end, time.max).replace(tzinfo=timezone.utc)
    if period == "month_current":
        start = today.replace(day=1)
        next_month = start.replace(year=start.year + 1, month=1) if start.month == 12 else start.replace(month=start.month + 1)
        end = next_month - timedelta(days=1)
        return datetime.combine(start, time.min).replace(tzinfo=timezone.utc), datetime.combine(end, time.max).replace(tzinfo=timezone.utc)
    if period == "month_prev":
        cur_start = today.replace(day=1)
        prev_end = cur_start - timedelta(days=1)
        prev_start = prev_end.replace(day=1)
        return datetime.combine(prev_start, time.min).replace(tzinfo=timezone.utc), datetime.combine(prev_end, time.max).replace(tzinfo=timezone.utc)
    return None, None


def normalize_call_type_filter(value: str | None) -> str | None:
    if not value:
        return None
    v = value.upper()
    # Canonical code in system is "КЦ" (Cyrillic).
    # Backward-compat: accept "KЦ"/"KC" from older scripts/UI.
    if v in {"KЦ", "КЦ", "KC"}:
        return "КЦ"
    return v


def parse_date_only(d: str | None, end_of_day: bool = False) -> datetime | None:
    if not d:
        return None
    day = datetime.strptime(d, "%Y-%m-%d").date()
    return datetime.combine(day, time.max if end_of_day else time.min).replace(tzinfo=timezone.utc)


def choose_date_range(period: str | None, date_from: str | None, date_to: str | None) -> tuple[datetime | None, datetime | None, str | None]:
    if period:
        start, end = resolve_period(period)
        return start, end, None

    custom_from = parse_date_only(date_from, end_of_day=False)
    custom_to = parse_date_only(date_to, end_of_day=True)
    if custom_from and custom_to and custom_to < custom_from:
        return custom_from, custom_to, "Дата 'по' не может быть раньше даты 'с'"
    return custom_from, custom_to, None


def build_calls_query(
    *,
    call_types: set[str],
    date_from: datetime | None,
    date_to: datetime | None,
    manager: str | None,
    status: str | None,
    call_type: str | None,
    transcription_detail: str = "none",
):
    """transcription_detail: 'none' — не грузить text (списки UI); 'full' — поле text для экспорта/API."""
    allowed = set(call_types)
    trans_opt = selectinload(Call.transcriptions)
    if transcription_detail != "full":
        trans_opt = trans_opt.load_only(
            Transcription.id,
            Transcription.call_id,
            Transcription.is_active,
            Transcription.model_name,
        )
    class_opt = selectinload(Call.classifications).load_only(
        CallClassification.id,
        CallClassification.call_id,
        CallClassification.is_active,
        CallClassification.topic_name,
        CallClassification.subtopic_name,
        CallClassification.confidence,
        CallClassification.reasoning,
    )
    query = (
        select(Call)
        .join(CallType, Call.call_type_id == CallType.id)
        .join(User, Call.manager_id == User.id)
        .options(
            selectinload(Call.call_type),
            selectinload(Call.manager),
            selectinload(Call.call_status),
            trans_opt,
            class_opt,
        )
        .where(CallType.code.in_(allowed))
    )

    ct = normalize_call_type_filter(call_type)
    if ct:
        query = query.where(CallType.code == ct)
    if date_from:
        query = query.where(Call.call_started_at >= date_from)
    if date_to:
        query = query.where(Call.call_started_at <= date_to)
    if manager:
        query = query.where(User.full_name == manager)
    if status:
        query = query.where(Call.status_id.in_(select(CallStatus.id).where(CallStatus.code == status)))
    return query


def users_department_for_role(role: str) -> str | None:
    if role == ROLE_911:
        return "911"
    if role == ROLE_KC:
        return "Contact Center"
    return None


def users_department_for_identity(identity: UiIdentity) -> str | None:
    has_kc = "КЦ" in identity.call_types
    has_911 = "911" in identity.call_types
    if has_kc and not has_911:
        return "Contact Center"
    if has_911 and not has_kc:
        return "911"
    return None


def get_active_transcription(call: Call):
    return next((t for t in call.transcriptions if t.is_active), None)


def get_active_classification(call: Call):
    return next((c for c in call.classifications if c.is_active), None)


def get_latest_summarization(call: Call):
    sums = getattr(call, "summarizations", None) or []
    if not sums:
        return None
    return max(sums, key=lambda s: s.id)


def safe_parts_count(call: Call) -> int:
    return getattr(call, "parts_count", 1)


def call_to_out(call: Call, include_text: bool) -> CallOut:
    active_trans = get_active_transcription(call)
    active_class = get_active_classification(call)
    latest_sum = get_latest_summarization(call)
    return CallOut(
        id=call.id,
        octell_call_id=call.octell_call_id,
        call_type=call.call_type.code,
        manager_folder=call.manager.manager_folder,
        manager_full_name=call.manager.full_name,
        manager_domain=call.manager.domain,
        status=call.call_status.code if call.call_status else "NEW",
        call_started_at=call.call_started_at,
        duration_seconds=call.duration_seconds,
        parts_count=safe_parts_count(call),
        transcription=active_trans.text if (active_trans and include_text) else None,
        has_transcription=active_trans is not None,
        topic=active_class.topic_name if active_class else None,
        subtopic=active_class.subtopic_name if active_class else None,
        classification_confidence=active_class.confidence if active_class else None,
        classification_reason=active_class.reasoning if active_class else None,
        summary_topic=latest_sum.topic if latest_sum else None,
        summary_outcome=latest_sum.outcome if latest_sum else None,
        summary_short=latest_sum.short_summary if latest_sum else None,
    )


@app.get("/", response_class=HTMLResponse)
def login_page(request: Request, error: str | None = Query(default=None)):
    if settings.ui_auth_mode == "trusted_headers":
        try:
            get_current_identity_ui(request)
            return RedirectResponse(url="/calls", status_code=302)
        except HTTPException:
            pass
        if settings.ui_superuser_enabled:
            return templates.TemplateResponse(
                "login.html",
                {
                    "request": request,
                    "show_header": False,
                    "active_page": "login",
                    "error": error,
                    "trusted_auth": False,
                    "superuser_via_trusted_mode": True,
                },
            )
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "show_header": False,
                "active_page": "login",
                "error": error or "Вход выполняется через корпоративный портал (нет заголовков доступа).",
                "trusted_auth": True,
                "superuser_via_trusted_mode": False,
            },
            status_code=403,
        )
    return templates.TemplateResponse(
        "login.html",
        {
            "request": request,
            "show_header": False,
            "active_page": "login",
            "error": error,
            "trusted_auth": False,
            "superuser_via_trusted_mode": False,
        },
    )


@app.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    try:
        ident = ui_login_authenticate(username=username.strip(), password=password)
    except HTTPException as exc:
        trusted_block = settings.ui_auth_mode == "trusted_headers" and not settings.ui_superuser_enabled
        return templates.TemplateResponse(
            "login.html",
            {
                "request": request,
                "show_header": False,
                "active_page": "login",
                "error": str(exc.detail) if isinstance(exc.detail, str) else "Ошибка входа",
                "trusted_auth": trusted_block,
                "superuser_via_trusted_mode": settings.ui_auth_mode == "trusted_headers" and settings.ui_superuser_enabled,
            },
            status_code=int(exc.status_code),
        )
    request.session["username"] = ident.username
    request.session["role"] = ident.role
    request.session["groups"] = ident.groups
    request.session["call_types"] = list(ident.call_types)
    request.session["catalog_access"] = ident.catalog_access
    request.session["pipeline_admin"] = ident.pipeline_admin
    return RedirectResponse(url="/calls", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)


@app.get("/ui/calls/{call_id}/active-transcription")
def ui_call_active_transcription(call_id: int, identity: UiIdentity = Depends(get_current_identity_ui)):
    """Текст активной транскрипции одним запросом (списки UI не грузят TEXT в общем select)."""
    allowed = identity.call_types
    db = SessionLocal()
    try:
        call_type_code = db.scalar(
            select(CallType.code).join(Call, Call.call_type_id == CallType.id).where(Call.id == call_id)
        )
        if call_type_code is None:
            raise HTTPException(status_code=404, detail="Call not found")
        if call_type_code not in allowed:
            raise HTTPException(status_code=403, detail="Forbidden")
        text = db.scalar(
            select(Transcription.text).where(Transcription.call_id == call_id, Transcription.is_active.is_(True))
        )
        return JSONResponse({"text": text})
    finally:
        db.close()


@app.get("/calls", response_class=HTMLResponse)
def calls_ui(
    request: Request,
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    status: str | None = Query(default=None),
    call_type: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
):
    role = identity.role
    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)

    db = SessionLocal()
    try:
        page_size = _UI_PAGE_SIZE
        query = build_calls_query(
            call_types=identity.call_types,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            status=status,
            call_type=call_type,
            transcription_detail="none",
        )
        total = int(db.scalar(select(func.count()).select_from(query.subquery())) or 0)
        if total > 0:
            max_offset = max(0, (total - 1) // page_size * page_size)
            if offset > max_offset:
                offset = max_offset
        calls = list(db.scalars(query.order_by(Call.call_started_at.desc()).offset(offset).limit(page_size)))
        rows = [call_to_out(c, include_text=False) for c in calls]
        ct_filter = normalize_call_type_filter(call_type) or ""
        list_volume_notice = total > _UI_LIST_SOFT_CAP
        pager = {
            "total": total,
            "offset": offset,
            "page_size": page_size,
            "from_row": offset + 1 if rows else 0,
            "to_row": offset + len(rows),
            "prev_href": f"/calls?{_calls_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, status=status, call_type=ct_filter or (call_type or ''), offset=max(0, offset - page_size))}"
            if offset > 0
            else None,
            "next_href": f"/calls?{_calls_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, status=status, call_type=ct_filter or (call_type or ''), offset=offset + page_size)}"
            if offset + len(rows) < total
            else None,
        }

        manager_stmt = (
            select(User.full_name)
            .join(Call, Call.manager_id == User.id)
            .join(CallType, Call.call_type_id == CallType.id)
            .where(CallType.code.in_(identity.call_types), User.full_name.is_not(None))
            .distinct()
            .order_by(User.full_name.asc())
        )
        managers = [m for m in db.scalars(manager_stmt) if m]

        return templates.TemplateResponse(
            "calls.html",
            {
                "request": request,
                "rows": rows,
                "role": role,
                "show_header": True,
                "active_page": "calls",
                "can_see_classification": identity_sees_kc_classification(identity),
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "managers": managers,
                "status_options": STATUS_OPTIONS,
                "date_error": date_error,
                "menu": menu_context("calls", identity),
                "filters": {
                    "period": period or "",
                    "date_from": date_from or "",
                    "date_to": date_to or "",
                    "manager": manager or "",
                    "status": status or "",
                    "call_type": normalize_call_type_filter(call_type) or "",
                },
                "pager": pager,
                "list_volume_notice": list_volume_notice,
                "list_soft_cap": _UI_LIST_SOFT_CAP,
            },
        )
    finally:
        db.close()


@app.get("/classified-calls", response_class=HTMLResponse)
def classified_calls_ui(
    request: Request,
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    subtopic: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
):
    role = identity.role
    if "КЦ" not in identity.call_types:
        raise HTTPException(status_code=403, detail="Нет доступа к классификации КЦ")

    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    db = SessionLocal()
    try:
        page_size = _UI_PAGE_SIZE
        role_types = identity.call_types & {"КЦ"}
        total = count_calls_with_active_classification(
            db,
            role_call_types=role_types,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            subtopic=subtopic,
        )
        if total > 0:
            max_offset = max(0, (total - 1) // page_size * page_size)
            if offset > max_offset:
                offset = max_offset
        calls = list_calls_with_active_classification(
            db,
            role_call_types=role_types,
            limit=page_size,
            offset=offset,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            subtopic=subtopic,
            load_transcription_text=False,
        )
        rows = [call_to_out(c, include_text=False) for c in calls]
        list_volume_notice = total > _UI_LIST_SOFT_CAP
        pager = {
            "total": total,
            "offset": offset,
            "page_size": page_size,
            "from_row": offset + 1 if rows else 0,
            "to_row": offset + len(rows),
            "prev_href": f"/classified-calls?{_classified_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, topic=topic, subtopic=subtopic, offset=max(0, offset - page_size))}"
            if offset > 0
            else None,
            "next_href": f"/classified-calls?{_classified_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, topic=topic, subtopic=subtopic, offset=offset + page_size)}"
            if offset + len(rows) < total
            else None,
        }

        # Filters UI (catalog-driven options)
        catalog_entries = list_topic_catalog_entries(db, include_inactive=False)
        topic_options = sorted({e.topic_name for e in catalog_entries})
        if topic:
            subtopic_options = sorted({e.subtopic_name for e in catalog_entries if e.topic_name == topic})
        else:
            subtopic_options = sorted({e.subtopic_name for e in catalog_entries})

        # Managers list (based on current date range only)
        manager_stmt = (
            select(User.full_name)
            .join(Call, Call.manager_id == User.id)
            .join(CallType, Call.call_type_id == CallType.id)
            .join(CallStatus, Call.status_id == CallStatus.id)
            .join(CallClassification, (CallClassification.call_id == Call.id) & (CallClassification.is_active.is_(True)))
            .where(
                CallStatus.code == "CLASSIFIED",
                CallType.code.in_(role_types),
            )
        )
        if final_from:
            manager_stmt = manager_stmt.where(Call.call_started_at >= final_from)
        if final_to:
            manager_stmt = manager_stmt.where(Call.call_started_at <= final_to)
        if topic:
            manager_stmt = manager_stmt.where(CallClassification.topic_name == topic)
        if subtopic:
            manager_stmt = manager_stmt.where(CallClassification.subtopic_name == subtopic)
        manager_stmt = manager_stmt.distinct().order_by(User.full_name.asc())
        managers = [m for m in db.scalars(manager_stmt) if m]
        return templates.TemplateResponse(
            "classified_calls.html",
            {
                "request": request,
                "rows": rows,
                "role": role,
                "show_header": True,
                "active_page": "classified_calls",
                "can_see_classification": True,
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "date_error": date_error,
                "menu": menu_context("classified_calls", identity),
                "filters": {
                    "period": period or "",
                    "date_from": date_from or "",
                    "date_to": date_to or "",
                    "manager": manager or "",
                    "topic": topic or "",
                    "subtopic": subtopic or "",
                },
                "pager": pager,
                "list_volume_notice": list_volume_notice,
                "list_soft_cap": _UI_LIST_SOFT_CAP,
                "managers": managers,
                "topic_options": topic_options,
                "subtopic_options": subtopic_options,
            },
        )
    finally:
        db.close()


@app.get("/summarized-calls", response_class=HTMLResponse)
def summarized_calls_ui(
    request: Request,
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    outcome: str | None = Query(default=None),
    offset: int = Query(default=0, ge=0),
):
    if not identity_sees_911_summarization(identity):
        raise HTTPException(status_code=403, detail="Нет доступа к саммаризации 911")

    role = identity.role
    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    db = SessionLocal()
    try:
        page_size = _UI_PAGE_SIZE
        total = count_calls_with_latest_summarization_911(
            db,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            outcome=outcome,
        )
        if total > 0:
            max_offset = max(0, (total - 1) // page_size * page_size)
            if offset > max_offset:
                offset = max_offset
        calls = list_calls_with_latest_summarization_911(
            db,
            limit=page_size,
            offset=offset,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            outcome=outcome,
            load_transcription_text=False,
        )
        rows = [call_to_out(c, include_text=False) for c in calls]
        list_volume_notice = total > _UI_LIST_SOFT_CAP
        pager = {
            "total": total,
            "offset": offset,
            "page_size": page_size,
            "from_row": offset + 1 if rows else 0,
            "to_row": offset + len(rows),
            "prev_href": f"/summarized-calls?{_summarized_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, topic=topic, outcome=outcome, offset=max(0, offset - page_size))}"
            if offset > 0
            else None,
            "next_href": f"/summarized-calls?{_summarized_list_querystring(period=period, date_from=date_from, date_to=date_to, manager=manager, topic=topic, outcome=outcome, offset=offset + page_size)}"
            if offset + len(rows) < total
            else None,
        }

        topic_options = list_distinct_latest_summary_topics_911(db)
        outcome_options = list_distinct_latest_summary_outcomes_911(db)
        managers = list_manager_names_for_summarized_911_filters(
            db,
            date_from=final_from,
            date_to=final_to,
            topic=topic,
            outcome=outcome,
        )

        return templates.TemplateResponse(
            "summarized_calls.html",
            {
                "request": request,
                "rows": rows,
                "role": role,
                "show_header": True,
                "active_page": "summarized_calls",
                "can_see_classification": identity_sees_kc_classification(identity),
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "date_error": date_error,
                "menu": menu_context("summarized_calls", identity),
                "filters": {
                    "period": period or "",
                    "date_from": date_from or "",
                    "date_to": date_to or "",
                    "manager": manager or "",
                    "topic": topic or "",
                    "outcome": outcome or "",
                },
                "pager": pager,
                "list_volume_notice": list_volume_notice,
                "list_soft_cap": _UI_LIST_SOFT_CAP,
                "managers": managers,
                "topic_options": topic_options,
                "outcome_options": outcome_options,
            },
        )
    finally:
        db.close()


@app.get("/users", response_class=HTMLResponse)
def users_ui(request: Request, identity: UiIdentity = Depends(get_current_identity_ui), limit: int = Query(default=1000, ge=1, le=5000)):
    role = identity.role
    db = SessionLocal()
    try:
        query = select(User).order_by(User.full_name.asc(), User.id.asc())
        dept = users_department_for_identity(identity)
        if dept:
            query = query.where(User.department == dept)
        rows = list(db.scalars(query.limit(limit)))
        return templates.TemplateResponse(
            "users.html",
            {
                "request": request,
                "rows": rows,
                "role": role,
                "show_header": True,
                "active_page": "users",
                "can_see_classification": identity_sees_kc_classification(identity),
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "menu": menu_context("users", identity),
            },
        )
    finally:
        db.close()


@app.get("/pipeline-runs", response_class=HTMLResponse)
def pipeline_runs_ui(request: Request, identity: UiIdentity = Depends(get_current_identity_ui), limit: int = Query(default=500, ge=1, le=2000)):
    if not identity.pipeline_admin:
        raise HTTPException(status_code=403, detail="Доступно только администраторам пайплайна")

    role = identity.role
    db = SessionLocal()
    try:
        runs = list(db.scalars(select(PipelineRun).order_by(PipelineRun.started_at.desc()).limit(limit)))
        return templates.TemplateResponse(
            "pipeline_runs.html",
            {
                "request": request,
                "rows": runs,
                "role": role,
                "show_header": True,
                "active_page": "pipeline_runs",
                "can_see_classification": identity_sees_kc_classification(identity),
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "menu": menu_context("pipeline_runs", identity),
            },
        )
    finally:
        db.close()


@app.get("/catalog", response_class=HTMLResponse)
def topic_catalog_ui(request: Request, identity: UiIdentity = Depends(get_current_identity_ui), include_inactive: bool = Query(default=True)):
    if not identity.catalog_access:
        raise HTTPException(status_code=403, detail="Нет доступа к справочнику")

    role = identity.role
    db = SessionLocal()
    try:
        rows = list_topic_catalog_entries(db, include_inactive=include_inactive)
        return templates.TemplateResponse(
            "catalog.html",
            {
                "request": request,
                "rows": rows,
                "role": role,
                "show_header": True,
                "active_page": "catalog",
                "can_see_classification": identity_sees_kc_classification(identity),
                "catalog_access": identity.catalog_access,
                "pipeline_admin": identity.pipeline_admin,
                "include_inactive": include_inactive,
                "menu": menu_context("catalog", identity),
            },
        )
    finally:
        db.close()


@app.post("/catalog/save")
def save_catalog_entry(
    entry_id: int | None = Form(default=None),
    topic_name: str = Form(...),
    subtopic_name: str = Form(...),
    description: str = Form(...),
    keywords_text: str = Form(...),
    synonyms_text: str | None = Form(default=None),
    negative_keywords_text: str | None = Form(default=None),
    is_active: str | None = Form(default="true"),
    identity: UiIdentity = Depends(get_current_identity_ui),
):
    if not identity.catalog_access:
        raise HTTPException(status_code=403, detail="Нет доступа к справочнику")

    topic_name = topic_name.strip()
    subtopic_name = subtopic_name.strip()
    description = description.strip()
    keywords_text = keywords_text.strip()
    if not topic_name or not subtopic_name or not description or not keywords_text:
        raise HTTPException(status_code=400, detail="Поля тема/подтема/описание/ключевые слова обязательны")

    db = SessionLocal()
    try:
        active_flag = str(is_active).lower() in {"true", "1", "on", "yes"}
        if entry_id:
            entry = update_topic_catalog_entry(
                db,
                entry_id=entry_id,
                topic_name=topic_name,
                subtopic_name=subtopic_name,
                description=description,
                keywords_text=keywords_text,
                synonyms_text=synonyms_text,
                negative_keywords_text=negative_keywords_text,
                is_active=active_flag,
            )
        else:
            entry = upsert_topic_catalog_entry(
                db,
                topic_name=topic_name,
                subtopic_name=subtopic_name,
                description=description,
                keywords_text=keywords_text,
                synonyms_text=synonyms_text,
                negative_keywords_text=negative_keywords_text,
                source_name="ui_admin",
                source_hash=entry_source_hash(topic_name, subtopic_name, description, keywords_text, synonyms_text),
                is_active=active_flag,
            )
        point_ids = sync_catalog_entries([entry])
        if point_ids:
            set_catalog_qdrant_point_id(db, entry.id, point_ids[0])
    finally:
        db.close()

    # Background enrichment: generate synonyms after UI change.
    # Enabled by default; can be disabled in prod via env to avoid unexpected GPU/CPU load.
    # (Tasks are routed to q.catalog by Celery config.)
    if os.getenv("CATALOG_AUTO_SYNONYMS", "1").strip().lower() in {"1", "true", "yes", "on"}:
        try:
            catalog_generate_synonyms.delay(entry.id)
        except Exception:
            # UI save must not fail if the queue is temporarily unavailable.
            pass

    return RedirectResponse(url="/catalog", status_code=303)


def _export_calls_excel_response(
    *,
    call_types: set[str],
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    status: str | None,
    call_type: str | None,
) -> StreamingResponse:
    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    if date_error:
        raise HTTPException(status_code=400, detail=date_error)

    db = SessionLocal()
    try:
        query = build_calls_query(
            call_types=call_types,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            status=status,
            call_type=call_type,
            transcription_detail="full",
        ).order_by(Call.call_started_at.desc())
        calls = list(db.scalars(query))

        wb = Workbook()
        ws = wb.active
        ws.title = "Calls"
        ws.append(["ID", "Octell ID", "Type", "Manager", "Manager Folder", "Status", "Started", "Duration", "Parts", "Topic", "Subtopic", "Transcription"])
        for c in calls:
            active_trans = get_active_transcription(c)
            active_class = get_active_classification(c)
            ws.append([
                c.id,
                c.octell_call_id,
                c.call_type.code,
                c.manager.full_name,
                c.manager.manager_folder,
                c.call_status.code if c.call_status else None,
                c.call_started_at.isoformat() if c.call_started_at else None,
                c.duration_seconds,
                safe_parts_count(c),
                active_class.topic_name if active_class else None,
                active_class.subtopic_name if active_class else None,
                active_trans.text if active_trans else None,
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="calls.xlsx"'},
        )
    finally:
        db.close()


@app.get("/api/calls/export.xlsx")
def export_calls_excel(
    role: str = Depends(get_current_role),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    status: str | None = Query(default=None),
    call_type: str | None = Query(default=None),
):
    return _export_calls_excel_response(
        call_types=allowed_call_types_for_role(role),
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        status=status,
        call_type=call_type,
    )


@app.get("/calls/export.xlsx")
def export_calls_excel_ui(
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    status: str | None = Query(default=None),
    call_type: str | None = Query(default=None),
):
    """Экспорт UI: типы звонков из сессии/заголовков (в т.ч. одновременно 911 и КЦ)."""
    return _export_calls_excel_response(
        call_types=identity.call_types,
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        status=status,
        call_type=call_type,
    )


def _export_classified_calls_excel_response(
    *,
    kc_call_types: set[str],
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    topic: str | None,
    subtopic: str | None,
) -> StreamingResponse:
    if not kc_call_types:
        raise HTTPException(status_code=403, detail="Нет доступа к классификации КЦ")

    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    if date_error:
        raise HTTPException(status_code=400, detail=date_error)

    db = SessionLocal()
    try:
        calls = list_calls_with_active_classification(
            db,
            role_call_types=kc_call_types,
            limit=200000,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            subtopic=subtopic,
            load_transcription_text=True,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Classified Calls"
        ws.append(["ID", "Octell ID", "Manager", "Started", "Duration", "Parts", "Topic", "Subtopic", "Confidence", "Reason", "Transcription"])

        for c in calls:
            active_trans = get_active_transcription(c)
            active_class = get_active_classification(c)
            ws.append([
                c.id,
                c.octell_call_id,
                c.manager.full_name if c.manager else None,
                c.call_started_at.isoformat() if c.call_started_at else None,
                c.duration_seconds,
                safe_parts_count(c),
                active_class.topic_name if active_class else None,
                active_class.subtopic_name if active_class else None,
                active_class.confidence if active_class else None,
                active_class.reasoning if active_class else None,
                active_trans.text if active_trans else None,
            ])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="classified_calls.xlsx"'},
        )
    finally:
        db.close()


@app.get("/api/classified-calls/export.xlsx")
def export_classified_calls_excel(
    role: str = Depends(get_current_role),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    subtopic: str | None = Query(default=None),
):
    kc_scope = allowed_call_types_for_role(role) & {"КЦ"}
    return _export_classified_calls_excel_response(
        kc_call_types=kc_scope,
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        topic=topic,
        subtopic=subtopic,
    )


@app.get("/classified-calls/export.xlsx")
def export_classified_calls_excel_ui(
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    subtopic: str | None = Query(default=None),
):
    kc_scope = identity.call_types & {"КЦ"}
    return _export_classified_calls_excel_response(
        kc_call_types=kc_scope,
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        topic=topic,
        subtopic=subtopic,
    )


def _export_summarized_911_excel_response(
    *,
    scope_911: bool,
    period: str | None,
    date_from: str | None,
    date_to: str | None,
    manager: str | None,
    topic: str | None,
    outcome: str | None,
) -> StreamingResponse:
    if not scope_911:
        raise HTTPException(status_code=403, detail="Нет доступа к саммаризации 911")

    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    if date_error:
        raise HTTPException(status_code=400, detail=date_error)

    db = SessionLocal()
    try:
        calls = list_calls_with_latest_summarization_911(
            db,
            limit=200_000,
            offset=0,
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            topic=topic,
            outcome=outcome,
            load_transcription_text=True,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Summarized 911"
        ws.append(
            [
                "ID",
                "Octell ID",
                "Manager",
                "Started",
                "Duration",
                "Parts",
                "Topic (summary)",
                "Outcome",
                "Short summary",
                "Transcription",
            ]
        )

        for c in calls:
            active_trans = get_active_transcription(c)
            latest_sum = get_latest_summarization(c)
            ws.append(
                [
                    c.id,
                    c.octell_call_id,
                    c.manager.full_name if c.manager else None,
                    c.call_started_at.isoformat() if c.call_started_at else None,
                    c.duration_seconds,
                    safe_parts_count(c),
                    latest_sum.topic if latest_sum else None,
                    latest_sum.outcome if latest_sum else None,
                    latest_sum.short_summary if latest_sum else None,
                    active_trans.text if active_trans else None,
                ]
            )

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="summarized_911_calls.xlsx"'},
        )
    finally:
        db.close()


@app.get("/api/summarized-calls/export.xlsx")
def export_summarized_calls_excel_api(
    role: str = Depends(get_current_role),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    outcome: str | None = Query(default=None),
):
    has_911 = bool(allowed_call_types_for_role(role) & {"911"})
    return _export_summarized_911_excel_response(
        scope_911=has_911,
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        topic=topic,
        outcome=outcome,
    )


@app.get("/summarized-calls/export.xlsx")
def export_summarized_calls_excel_ui(
    identity: UiIdentity = Depends(get_current_identity_ui),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    topic: str | None = Query(default=None),
    outcome: str | None = Query(default=None),
):
    return _export_summarized_911_excel_response(
        scope_911=identity_sees_911_summarization(identity),
        period=period,
        date_from=date_from,
        date_to=date_to,
        manager=manager,
        topic=topic,
        outcome=outcome,
    )


@app.get("/api/calls", response_model=CallsResponse)
def calls_api(
    role: str = Depends(get_current_role),
    period: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    manager: str | None = Query(default=None),
    status: str | None = Query(default=None),
    call_type: str | None = Query(default=None),
    limit: int = Query(default=1000, ge=1, le=_API_LIMIT_MAX),
    offset: int = Query(default=0, ge=0),
    include_text: bool = Query(default=False),
):
    final_from, final_to, date_error = choose_date_range(period, date_from, date_to)
    if date_error:
        raise HTTPException(status_code=400, detail=date_error)

    db = SessionLocal()
    try:
        t_detail = "full" if include_text else "none"
        query = build_calls_query(
            call_types=allowed_call_types_for_role(role),
            date_from=final_from,
            date_to=final_to,
            manager=manager,
            status=status,
            call_type=call_type,
            transcription_detail=t_detail,
        )
        total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
        calls = list(db.scalars(query.order_by(Call.call_started_at.desc()).offset(offset).limit(limit)))
        return CallsResponse(items=[call_to_out(c, include_text=include_text) for c in calls], total=int(total))
    finally:
        db.close()


@app.get("/api/users", response_model=UsersResponse)
def users_api(role: str = Depends(get_current_role), limit: int = Query(default=1000, ge=1, le=5000), offset: int = Query(default=0, ge=0)):
    db = SessionLocal()
    try:
        query = select(User)
        dept = users_department_for_role(role)
        if dept:
            query = query.where(User.department == dept)
        total = db.scalar(select(func.count()).select_from(query.subquery())) or 0
        rows = list(db.scalars(query.order_by(User.full_name.asc(), User.id.asc()).offset(offset).limit(limit)))
        return UsersResponse(items=[UserOut(id=r.id, full_name=r.full_name, domain=r.domain, department=r.department) for r in rows], total=int(total))
    finally:
        db.close()


@app.get("/api/pipeline-runs", response_model=PipelineRunsResponse)
def pipeline_runs_api(role: str = Depends(get_current_role), limit: int = Query(default=500, ge=1, le=2000), offset: int = Query(default=0, ge=0)):
    if role != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Доступно только ADMIN")

    db = SessionLocal()
    try:
        total = db.scalar(select(func.count()).select_from(PipelineRun)) or 0
        rows = list(db.scalars(select(PipelineRun).order_by(PipelineRun.started_at.desc()).offset(offset).limit(limit)))
        return PipelineRunsResponse(
            items=[
                PipelineRunOut(
                    id=r.id,
                    pipeline_code=r.pipeline_code,
                    status=r.status,
                    started_at=r.started_at,
                    finished_at=r.finished_at,
                    duration_seconds=r.duration_seconds,
                    processed_calls=r.processed_calls,
                    total_audio_seconds=r.total_audio_seconds,
                    avg_rtf=r.avg_rtf,
                    error_message=r.error_message,
                )
                for r in rows
            ],
            total=int(total),
        )
    finally:
        db.close()


@app.get("/api/catalog", response_model=TopicCatalogEntriesResponse)
def catalog_api(role: str = Depends(get_current_role), include_inactive: bool = Query(default=True)):
    if role != ROLE_ADMIN:
        raise HTTPException(status_code=403, detail="Доступно только ADMIN")
    db = SessionLocal()
    try:
        rows = list_topic_catalog_entries(db, include_inactive=include_inactive)
        return TopicCatalogEntriesResponse(
            items=[
                TopicCatalogEntryOut(
                    id=row.id,
                    topic_name=row.topic_name,
                    subtopic_name=row.subtopic_name,
                    description=row.description,
                    keywords_text=row.keywords_text,
                    synonyms_text=row.synonyms_text,
                    negative_keywords_text=row.negative_keywords_text,
                    is_active=row.is_active,
                    qdrant_point_id=row.qdrant_point_id,
                    updated_at=row.updated_at,
                )
                for row in rows
            ],
            total=len(rows),
        )
    finally:
        db.close()