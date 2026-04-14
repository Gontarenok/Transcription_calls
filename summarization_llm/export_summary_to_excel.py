import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT_DIR = "output_summary"
REPORT_DIR = "output_reports"
AUDIO_DIR = "output_audio"
INPUT_DIR = "C:/Audio_share/Night"


def clean_value(value: str) -> str:
    """Чистим значение от мусора: *, лишние пробелы, дефисы и точки"""
    value = value.strip()
    value = value.replace("*", "")  # убираем все *
    value = re.sub(r"\s+", " ", value)  # схлопываем пробелы
    value = re.sub(r"[-–—]+\s*$", "", value)  # убираем хвостовые дефисы
    value = re.sub(r"[.\s]+$", "", value)  # убираем точки/пробелы в конце
    return value.strip()


def parse_summary(text: str) -> dict:
    """
    Разбирает текст саммари по ключевым полям.
    Возвращает словарь с колонками: Участники, Платформа, Суть, Действие, Итог, Краткое саммари
    """
    fields = ["Участники", "Платформа", "Тема", "Суть", "Действие в результате диалога", "Итог", "Краткое саммари"]
    result = {f: "Не указано" for f in fields}

    for i, field in enumerate(fields):
        pattern = rf"{field}:\s*(.*?)(?=(?:{'|'.join(fields[i + 1:])}):|$)"
        match = re.search(pattern, text, re.S | re.IGNORECASE)
        if match:
            result[field] = clean_value(match.group(1))
    return result


def format_excel(report_path: str):
    """Оформляем Excel: автоширина, жирный заголовок, границы"""
    wb = load_workbook(report_path)
    ws = wb.active

    # стили
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # синий фон для заголовков
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # заголовки
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    wb.save(report_path)


def get_latest_subdir(base_dir: str) -> str:
    """Возвращает путь к последней (по дате модификации) подпапке"""
    subdirs = [os.path.join(base_dir, d) for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]
    if not subdirs:
        raise SystemExit(f"❌ Нет папок в {base_dir}")
    return max(subdirs, key=os.path.getmtime)


def load_transcription(file_name: str) -> str:
    """
    Загружает исходную транскрибацию по имени файла summary.
    file_name: например mix_..._summary.txt -> ищем mix_...txt в последней папке output_audio
    """
    latest_audio_dir = get_latest_subdir(AUDIO_DIR)
    base_name = file_name.replace("_summary.txt", ".txt")
    audio_trans_path = os.path.join(latest_audio_dir, base_name)

    transcription = "⚠️ Файл транскрибации не найден"
    if os.path.exists(audio_trans_path):
        with open(audio_trans_path, encoding="utf8") as f:
            transcription = f.read().strip()

    # ищем исходный звонок
    base_call_name = file_name.replace("_summary.txt", "")
    call_path = None
    for ext in [".mp3", ".wav", ".m4a"]:
        candidate = os.path.join(INPUT_DIR, base_call_name + ext)
        if os.path.exists(candidate):
            call_path = candidate
            break

    call_date, call_time = "", ""
    if call_path:
        mtime = os.path.getmtime(call_path)
        dt = datetime.fromtimestamp(mtime)
        call_date = dt.strftime("%Y-%m-%d")
        call_time = dt.strftime("%H:%M:%S")
    else:
        transcription += " ⚠️ Исходный звонок не найден"

    return transcription, call_date, call_time


def main():
    latest_dir = get_latest_subdir(OUTPUT_DIR)
    print(f"📂 Используем папку с саммари: {latest_dir}")

    files = sorted([f for f in os.listdir(latest_dir) if f.endswith("_summary.txt")])
    if not files:
        raise SystemExit("❌ Нет файлов с саммари в выбранной папке")

    rows = []
    for file in files:
        path = os.path.join(latest_dir, file)
        with open(path, encoding="utf8") as f:
            text = f.read()

        parsed = parse_summary(text)
        parsed["Файл"] = file

        transcription, call_date, call_time = load_transcription(file)
        parsed["Транскрибация"] = transcription
        parsed["Дата звонка"] = call_date
        parsed["Время звонка"] = call_time

        rows.append(parsed)

    df = pd.DataFrame(
        rows,
        columns=[
            "Файл", "Дата звонка", "Время звонка",
            "Участники", "Платформа", "Тема", "Суть",
            "Действие в результате диалога", "Итог", "Краткое саммари", "Транскрибация"
        ]
    )

    os.makedirs(REPORT_DIR, exist_ok=True)
    report_name = f"calls_report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    report_path = os.path.join(REPORT_DIR, report_name)
    df.to_excel(report_path, index=False)

    format_excel(report_path)
    print(f"✅ Отчёт сохранён и отформатирован: {report_path}")


if __name__ == "__main__":
    main()
