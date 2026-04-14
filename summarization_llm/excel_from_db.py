"""Excel-отчёт по звонкам 911 из БД (транскрипт + поля саммари)."""

from __future__ import annotations

from datetime import timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from db.models import Call

from summarization_llm.weekly_stats import latest_summarization


def _active_transcription_text(call: Call) -> str:
    for t in call.transcriptions:
        if t.is_active:
            return (t.text or "").strip()
    return ""


def calls_to_excel_rows(calls: list[Call]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for call in calls:
        summ = latest_summarization(call)
        started = call.call_started_at
        if started.tzinfo:
            started = started.astimezone(timezone.utc)
        rows.append(
            {
                "Файл": call.file_name,
                "Дата звонка": started.strftime("%Y-%m-%d"),
                "Время звонка": started.strftime("%H:%M:%S"),
                "Участники": (summ.participants if summ else None) or "Не указано",
                "Платформа": (summ.platform if summ else None) or "Не указано",
                "Тема": (summ.topic if summ else None) or "Не указано",
                "Суть": (summ.essence if summ else None) or "Не указано",
                "Действие в результате диалога": (summ.action_result if summ else None) or "Не указано",
                "Итог": (summ.outcome if summ else None) or "Не указано",
                "Краткое саммари": (summ.short_summary if summ else None) or "Не указано",
                "Транскрибация": _active_transcription_text(call) or "—",
            }
        )
    return rows


def format_excel_report(report_path: str | Path) -> None:
    path = Path(report_path)
    wb = load_workbook(path)
    ws = wb.active
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    wb.save(path)


def export_911_calls_to_excel(calls: list[Call], output_path: str | Path) -> Path:
    """Сохраняет отчёт и возвращает путь к файлу."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = calls_to_excel_rows(calls)
    columns = [
        "Файл",
        "Дата звонка",
        "Время звонка",
        "Участники",
        "Платформа",
        "Тема",
        "Суть",
        "Действие в результате диалога",
        "Итог",
        "Краткое саммари",
        "Транскрибация",
    ]
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(path, index=False)
    format_excel_report(path)
    return path
